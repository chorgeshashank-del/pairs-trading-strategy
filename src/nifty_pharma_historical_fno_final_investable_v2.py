import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, date
import hashlib
import re
import shutil
import sys

try:
    import openpyxl
except ImportError as exc:
    raise ImportError(
        "openpyxl is required. Install with: pip install openpyxl"
    ) from exc


# =============================================================================
# NIFTY PHARMA — HISTORICAL F&O ELIGIBILITY + FINAL INVESTABLE UNIVERSE
# ROBUST V2 — DOES NOT DEPEND ON A SPECIFIC NSE WORKBOOK HEADER LAYOUT
# =============================================================================
#
# WHY V2 EXISTS
# -------------
# The first version assumed the official NSE workbook had a recognizable
# single-row header. The downloaded workbook has a different layout.
#
# This version instead:
#   1. reads the already-downloaded official NSE workbook cell-by-cell
#   2. searches for OUR research-universe symbols directly
#   3. infers INTRODUCTION / EXCLUSION date columns from:
#        - sheet name
#        - multi-row header text above each column
#        - nearby heading text
#   4. supports separate Introduction and Exclusion sheets
#   5. supports combined sheets with both dates
#   6. builds historical F&O episodes chronologically
#   7. never guesses an ambiguous case
#
# It then applies:
#   PRE_LIQUIDITY_ELIGIBLE
#   AND median daily traded value >= Rs 10 crore
#   AND confirmed historical F&O eligibility
#
# No pair selection or strategy P&L is used.
# =============================================================================


# =============================================================================
# 1. PATHS
# =============================================================================

PROJECT_ROOT = Path(r"C:\fin proj")

FORMATION_DIR = (
    PROJECT_ROOT
    / "nse_pharma_formation_investability"
)

FORMATION_DIAGNOSTICS_FILE = (
    FORMATION_DIR
    / "02_STOCK_FORMATION_DIAGNOSTICS.csv"
)

OUTPUT_DIR = (
    PROJECT_ROOT
    / "nse_pharma_final_investable_universe"
)

RAW_DIR = (
    OUTPUT_DIR
    / "raw_nse_fno_tracker"
)

OUTPUT_DIR.mkdir(
    parents=True,
    exist_ok=True
)

RAW_DIR.mkdir(
    parents=True,
    exist_ok=True
)


TRACKER_AS_OF = pd.Timestamp("2026-07-24")

TRACKER_FILENAME = (
    "FO_Stock_Introduction_and_Exclusion_Tracker_24-07-2026_"
    "20260724111707.xlsx"
)

TRACKER_LOCAL_FILE = (
    RAW_DIR
    / TRACKER_FILENAME
)


LIQUIDITY_THRESHOLD_RUPEES = 100_000_000.0  # Rs 10 crore


# =============================================================================
# 2. SYMBOL NORMALIZATION
# =============================================================================

SYMBOL_ALIAS_TO_COMPANY_ID = {
    "CADILAHC": "ZYDUS",
    "ZYDUSLIFE": "ZYDUS",
    "AJANTAPHARM": "AJANTPHARM",
}


def clean_symbol(value):
    if value is None:
        return None

    if isinstance(value, float) and np.isnan(value):
        return None

    text = str(value).strip().upper()

    if text in {
        "",
        "NAN",
        "NONE",
        "NA",
        "N/A",
        "-",
        "--",
    }:
        return None

    text = re.sub(
        r"\s+",
        "",
        text
    )

    return text


def symbol_to_company_id(symbol):
    symbol = clean_symbol(
        symbol
    )

    if symbol is None:
        return None

    return SYMBOL_ALIAS_TO_COMPANY_ID.get(
        symbol,
        symbol
    )


def normalize_text(value):
    if value is None:
        return ""

    if isinstance(value, float) and np.isnan(value):
        return ""

    text = str(value).upper()

    text = re.sub(
        r"[\r\n\t]+",
        " ",
        text
    )

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip()


def compact_text(value):
    text = normalize_text(
        value
    )

    return re.sub(
        r"[^A-Z0-9]+",
        "_",
        text
    ).strip("_")


# =============================================================================
# 3. DATE HELPERS
# =============================================================================

def parse_date_value(value):
    if value is None:
        return pd.NaT

    if isinstance(
        value,
        (
            pd.Timestamp,
            datetime,
            date,
        )
    ):
        return pd.Timestamp(
            value
        ).normalize()

    if isinstance(
        value,
        (
            int,
            float,
            np.integer,
            np.floating,
        )
    ):
        if pd.isna(value):
            return pd.NaT

        # Excel serial date range.
        if 20_000 <= float(value) <= 80_000:
            return (
                pd.Timestamp("1899-12-30")
                +
                pd.to_timedelta(
                    float(value),
                    unit="D"
                )
            ).normalize()

    text = str(value).strip()

    if text.upper() in {
        "",
        "NAN",
        "NONE",
        "NA",
        "N/A",
        "-",
        "--",
        "NOT APPLICABLE",
    }:
        return pd.NaT

    # Do not turn pure serial-looking non-date identifiers into dates.
    if re.fullmatch(
        r"\d{1,5}",
        text
    ):
        number = int(
            text
        )

        if 20_000 <= number <= 80_000:
            return (
                pd.Timestamp("1899-12-30")
                +
                pd.to_timedelta(
                    number,
                    unit="D"
                )
            ).normalize()

        return pd.NaT

    parsed = pd.to_datetime(
        text,
        dayfirst=True,
        errors="coerce"
    )

    if pd.isna(
        parsed
    ):
        return pd.NaT

    parsed = pd.Timestamp(
        parsed
    ).normalize()

    # Plausibility bound for this project.
    if not (
        pd.Timestamp("1995-01-01")
        <= parsed
        <= pd.Timestamp("2030-12-31")
    ):
        return pd.NaT

    return parsed


def to_bool(series):
    if series.dtype == bool:
        return series

    return (
        series
        .astype(str)
        .str.strip()
        .str.upper()
        .isin(
            [
                "TRUE",
                "1",
                "YES",
                "Y",
            ]
        )
    )


def sha256_file(path):
    h = hashlib.sha256()

    with open(
        path,
        "rb"
    ) as f:

        while True:

            chunk = f.read(
                1024 * 1024
            )

            if not chunk:
                break

            h.update(
                chunk
            )

    return h.hexdigest()


# =============================================================================
# 4. LOCATE THE OFFICIAL TRACKER
# =============================================================================

def locate_tracker():
    """
    Reuse the official workbook already downloaded by the first program.

    Search only under this project's final-universe/raw directory first.
    If necessary, search C:\fin proj for the exact frozen filename.
    """

    if TRACKER_LOCAL_FILE.exists():
        return TRACKER_LOCAL_FILE

    matches = list(
        PROJECT_ROOT.rglob(
            TRACKER_FILENAME
        )
    )

    if len(
        matches
    ) == 1:

        # Copy into the frozen raw location.
        shutil.copy2(
            matches[0],
            TRACKER_LOCAL_FILE
        )

        return TRACKER_LOCAL_FILE

    if len(
        matches
    ) == 0:

        raise FileNotFoundError(
            "\nThe official NSE F&O tracker is not present locally.\n"
            "The previous program already downloaded it, so it should exist.\n\n"
            f"Expected filename:\n{TRACKER_FILENAME}\n\n"
            f"Searched under:\n{PROJECT_ROOT}"
        )

    # If several exact copies exist, compare hashes rather than guessing.
    hashes = {
        sha256_file(
            p
        ):
            []
        for p in matches
    }

    for p in matches:
        hashes[
            sha256_file(
                p
            )
        ].append(
            p
        )

    if len(
        hashes
    ) == 1:

        chosen = matches[0]

        shutil.copy2(
            chosen,
            TRACKER_LOCAL_FILE
        )

        return TRACKER_LOCAL_FILE

    raise RuntimeError(
        "\nMultiple DIFFERENT files with the frozen tracker filename exist.\n"
        "Refusing to guess which workbook is authoritative.\n\n"
        +
        "\n".join(
            str(p)
            for p in matches
        )
    )


# =============================================================================
# 5. LOAD FORMATION DIAGNOSTICS FIRST
# =============================================================================

if not FORMATION_DIAGNOSTICS_FILE.exists():

    raise FileNotFoundError(
        "\nFormation diagnostics not found:\n"
        f"{FORMATION_DIAGNOSTICS_FILE}"
    )


diagnostics = pd.read_csv(
    FORMATION_DIAGNOSTICS_FILE,
    low_memory=False
)

diagnostics.columns = [
    compact_text(
        col
    )
    for col in diagnostics.columns
]


required_diag = {
    "FORMATION_DATE",
    "COMPANY_ID",
    "IN_NIFTY_PHARMA_AT_FORMATION",
    "PRE_LIQUIDITY_ELIGIBLE",
    "MEDIAN_DAILY_TRADED_VALUE",
    "BLOCK_TYPE",
}

missing_diag = (
    required_diag
    -
    set(
        diagnostics.columns
    )
)

if missing_diag:

    raise ValueError(
        "\nFormation diagnostics missing required columns:\n"
        f"{sorted(missing_diag)}"
    )


diagnostics[
    "FORMATION_DATE"
] = pd.to_datetime(
    diagnostics[
        "FORMATION_DATE"
    ],
    errors="coerce"
)

if diagnostics[
    "FORMATION_DATE"
].isna().any():

    raise ValueError(
        "Invalid FORMATION_DATE in formation diagnostics."
    )


diagnostics[
    "COMPANY_ID"
] = diagnostics[
    "COMPANY_ID"
].map(
    clean_symbol
)


diagnostics[
    "IN_NIFTY_PHARMA_AT_FORMATION"
] = to_bool(
    diagnostics[
        "IN_NIFTY_PHARMA_AT_FORMATION"
    ]
)

diagnostics[
    "PRE_LIQUIDITY_ELIGIBLE"
] = to_bool(
    diagnostics[
        "PRE_LIQUIDITY_ELIGIBLE"
    ]
)


diagnostics[
    "MEDIAN_DAILY_TRADED_VALUE"
] = pd.to_numeric(
    diagnostics[
        "MEDIAN_DAILY_TRADED_VALUE"
    ],
    errors="coerce"
)


if diagnostics[
    [
        "FORMATION_DATE",
        "COMPANY_ID",
    ]
].duplicated().any():

    raise RuntimeError(
        "Duplicate FORMATION_DATE + COMPANY_ID rows in diagnostics."
    )


# Build the exact set of historical symbols we need to recognize in the NSE
# workbook. This is far more robust than trying to parse every tracker stock.
research_company_ids = set(
    diagnostics[
        "COMPANY_ID"
    ].dropna().unique()
)


research_symbols = set(
    research_company_ids
)


for col in [
    "SYMBOL_ON_FORMATION_DATE",
    "MEMBERSHIP_SYMBOL",
]:

    if col in diagnostics.columns:

        research_symbols.update(
            clean_symbol(
                x
            )
            for x in diagnostics[
                col
            ].dropna().unique()
            if clean_symbol(
                x
            ) is not None
        )


# Add known historical aliases explicitly.
research_symbols.update(
    {
        "CADILAHC",
        "ZYDUSLIFE",
        "AJANTAPHARM",
        "AJANTPHARM",
    }
)


# Remove accidental stable ID that may not itself be an NSE ticker.
research_symbols.discard(
    "ZYDUS"
)


# =============================================================================
# 6. ROBUST WORKBOOK PARSER
# =============================================================================

INTRO_WORDS = [
    "INTRODUCTION",
    "INTRODUCED",
    "INTRO",
    "INCLUSION",
    "INCLUDED",
    "ENTRY",
    "ADDITION",
    "ADDED",
    "INDUCTION",
    "INDUCTED",
    "COMMENCEMENT",
    "COMMENCE",
]

EXCLUSION_WORDS = [
    "EXCLUSION",
    "EXCLUDED",
    "EXCLUDE",
    "EXIT",
    "REMOVAL",
    "REMOVED",
    "DISCONTINUATION",
    "DISCONTINUED",
    "WITHDRAWAL",
    "WITHDRAWN",
]

DATE_WORDS = [
    "DATE",
    "W.E.F",
    "WEF",
    "EFFECTIVE",
    "EFFECT",
    "FROM",
]


def semantic_from_text(text):
    text = normalize_text(
        text
    )

    has_intro = any(
        word in text
        for word in INTRO_WORDS
    )

    has_excl = any(
        word in text
        for word in EXCLUSION_WORDS
    )

    if (
        has_intro
        and
        not has_excl
    ):
        return "INTRO"

    if (
        has_excl
        and
        not has_intro
    ):
        return "EXCLUSION"

    if (
        has_intro
        and
        has_excl
    ):
        return "BOTH"

    return "UNKNOWN"


def sheet_semantic(sheet_name):
    return semantic_from_text(
        sheet_name
    )


def build_column_context(
    ws,
    row_num,
    col_num,
    lookback=12
):
    """
    Concatenate non-empty cells ABOVE the symbol row in this date column and
    neighbouring columns. This handles multi-row / merged-style headings.
    """

    texts = []

    r_start = max(
        1,
        row_num
        -
        lookback
    )

    c_start = max(
        1,
        col_num
        -
        1
    )

    c_end = min(
        ws.max_column,
        col_num
        +
        1
    )

    for r in range(
        r_start,
        row_num
    ):

        for c in range(
            c_start,
            c_end
            +
            1
        ):

            value = ws.cell(
                r,
                c
            ).value

            text = normalize_text(
                value
            )

            if text:
                texts.append(
                    text
                )

    return " | ".join(
        texts
    )


def row_context(
    ws,
    row_num,
    lookback=5
):
    texts = []

    r_start = max(
        1,
        row_num
        -
        lookback
    )

    for r in range(
        r_start,
        row_num
    ):

        for c in range(
            1,
            ws.max_column
            +
            1
        ):

            text = normalize_text(
                ws.cell(
                    r,
                    c
                ).value
            )

            if text:
                texts.append(
                    text
                )

    return " | ".join(
        texts
    )


def identify_date_semantic(
    ws,
    row_num,
    col_num
):
    col_context = build_column_context(
        ws,
        row_num,
        col_num,
        lookback=15
    )

    sem = semantic_from_text(
        col_context
    )

    if sem in {
        "INTRO",
        "EXCLUSION",
    }:
        return (
            sem,
            "COLUMN_HEADER_CONTEXT",
            col_context
        )

    sheet_sem = sheet_semantic(
        ws.title
    )

    if sheet_sem in {
        "INTRO",
        "EXCLUSION",
    }:
        return (
            sheet_sem,
            "SHEET_NAME",
            ws.title
        )

    broader = (
        row_context(
            ws,
            row_num,
            lookback=8
        )
        +
        " | "
        +
        ws.title
    )

    broader_sem = semantic_from_text(
        broader
    )

    if broader_sem in {
        "INTRO",
        "EXCLUSION",
    }:
        return (
            broader_sem,
            "NEARBY_HEADING_CONTEXT",
            broader
        )

    return (
        "UNKNOWN",
        "UNRESOLVED",
        col_context
    )


def scan_tracker_for_research_symbols(
    workbook_path
):
    wb = openpyxl.load_workbook(
        workbook_path,
        data_only=True,
        read_only=False
    )

    symbol_occurrences = []
    event_records = []
    unresolved_records = []


    for ws in wb.worksheets:

        print(
            f"Scanning sheet: {ws.title}"
        )

        # First pass: exact occurrences of one of our known historical symbols.
        for row in ws.iter_rows():

            for cell in row:

                value_symbol = clean_symbol(
                    cell.value
                )

                if (
                    value_symbol is None
                    or
                    value_symbol not in research_symbols
                ):
                    continue

                row_num = cell.row
                symbol_col = cell.column

                company_id = symbol_to_company_id(
                    value_symbol
                )

                symbol_occurrences.append(
                    {
                        "SHEET":
                            ws.title,

                        "ROW":
                            row_num,

                        "SYMBOL_COLUMN":
                            symbol_col,

                        "TRACKER_SYMBOL":
                            value_symbol,

                        "COMPANY_ID":
                            company_id,
                    }
                )


                # Collect every plausible date in the SAME data row.
                row_dates = []

                for c in range(
                    1,
                    ws.max_column
                    +
                    1
                ):

                    if c == symbol_col:
                        continue

                    raw_value = ws.cell(
                        row_num,
                        c
                    ).value

                    parsed_date = parse_date_value(
                        raw_value
                    )

                    if pd.isna(
                        parsed_date
                    ):
                        continue

                    # Avoid Excel-created dates from irrelevant ancient/future
                    # numeric cells; our parser already uses plausibility bound.
                    semantic, semantic_source, context = (
                        identify_date_semantic(
                            ws,
                            row_num,
                            c
                        )
                    )

                    row_dates.append(
                        {
                            "DATE_COLUMN":
                                c,

                            "RAW_DATE_VALUE":
                                raw_value,

                            "EVENT_DATE":
                                parsed_date,

                            "EVENT_TYPE":
                                semantic,

                            "SEMANTIC_SOURCE":
                                semantic_source,

                            "HEADER_CONTEXT":
                                context,
                        }
                    )


                # Deduplicate identical date/column interpretations.
                unique_dates = []

                seen = set()

                for item in row_dates:

                    key = (
                        item[
                            "DATE_COLUMN"
                        ],
                        item[
                            "EVENT_DATE"
                        ],
                        item[
                            "EVENT_TYPE"
                        ],
                    )

                    if key in seen:
                        continue

                    seen.add(
                        key
                    )

                    unique_dates.append(
                        item
                    )


                resolved = [
                    x
                    for x in unique_dates
                    if x[
                        "EVENT_TYPE"
                    ] in {
                        "INTRO",
                        "EXCLUSION",
                    }
                ]


                # If a sheet clearly represents one event type, and there is
                # exactly one date in the row, use the sheet semantic.
                if (
                    len(
                        resolved
                    ) == 0
                    and
                    len(
                        unique_dates
                    ) == 1
                ):

                    sh_sem = sheet_semantic(
                        ws.title
                    )

                    if sh_sem in {
                        "INTRO",
                        "EXCLUSION",
                    }:

                        unique_dates[
                            0
                        ][
                            "EVENT_TYPE"
                        ] = sh_sem

                        unique_dates[
                            0
                        ][
                            "SEMANTIC_SOURCE"
                        ] = "SHEET_NAME_SINGLE_DATE_FALLBACK"

                        resolved = [
                            unique_dates[
                                0
                            ]
                        ]


                # Save resolved events.
                for item in resolved:

                    event_records.append(
                        {
                            "SHEET":
                                ws.title,

                            "ROW":
                                row_num,

                            "TRACKER_SYMBOL":
                                value_symbol,

                            "COMPANY_ID":
                                company_id,

                            "EVENT_TYPE":
                                item[
                                    "EVENT_TYPE"
                                ],

                            "EVENT_DATE":
                                item[
                                    "EVENT_DATE"
                                ],

                            "DATE_COLUMN":
                                item[
                                    "DATE_COLUMN"
                                ],

                            "RAW_DATE_VALUE":
                                item[
                                    "RAW_DATE_VALUE"
                                ],

                            "SEMANTIC_SOURCE":
                                item[
                                    "SEMANTIC_SOURCE"
                                ],

                            "HEADER_CONTEXT":
                                item[
                                    "HEADER_CONTEXT"
                                ],
                        }
                    )


                # Anything still unresolved is written for review rather than
                # silently interpreted.
                unresolved = [
                    x
                    for x in unique_dates
                    if x[
                        "EVENT_TYPE"
                    ] not in {
                        "INTRO",
                        "EXCLUSION",
                    }
                ]

                if unresolved:

                    unresolved_records.append(
                        {
                            "SHEET":
                                ws.title,

                            "ROW":
                                row_num,

                            "TRACKER_SYMBOL":
                                value_symbol,

                            "COMPANY_ID":
                                company_id,

                            "UNRESOLVED_DATE_COUNT":
                                len(
                                    unresolved
                                ),

                            "UNRESOLVED_DATES":
                                ";".join(
                                    str(
                                        x[
                                            "EVENT_DATE"
                                        ].date()
                                    )
                                    for x in unresolved
                                ),

                            "SHEET_SEMANTIC":
                                sheet_semantic(
                                    ws.title
                                ),

                            "ROW_CONTEXT":
                                row_context(
                                    ws,
                                    row_num,
                                    lookback=8
                                ),
                        }
                    )


    occurrences_df = pd.DataFrame(
        symbol_occurrences
    )

    events_df = pd.DataFrame(
        event_records
    )

    unresolved_df = pd.DataFrame(
        unresolved_records
    )


    if not events_df.empty:

        events_df = events_df.drop_duplicates(
            subset=[
                "TRACKER_SYMBOL",
                "COMPANY_ID",
                "EVENT_TYPE",
                "EVENT_DATE",
            ],
            keep="first"
        ).sort_values(
            [
                "COMPANY_ID",
                "EVENT_DATE",
                "EVENT_TYPE",
            ]
        ).reset_index(
            drop=True
        )


    return (
        occurrences_df,
        events_df,
        unresolved_df,
        wb.sheetnames,
    )


# =============================================================================
# 7. BUILD F&O EPISODES FROM INTRODUCTION / EXCLUSION EVENTS
# =============================================================================

def build_episodes(
    events_df
):
    """
    Construct active intervals without guessing.

    For each COMPANY_ID:
      - sort INTRO and EXCLUSION events
      - each INTRO opens an episode
      - next EXCLUSION after that INTRO closes it
      - exclusion date is the first NON-eligible date

    Any chronology problem is put into an ambiguity table.
    """

    episodes = []
    ambiguities = []


    if events_df.empty:

        return (
            pd.DataFrame(),
            pd.DataFrame(
                [
                    {
                        "COMPANY_ID":
                            "",
                        "ISSUE":
                            "NO_EVENTS_PARSED",
                    }
                ]
            )
        )


    for company_id, grp in events_df.groupby(
        "COMPANY_ID"
    ):

        g = grp.sort_values(
            [
                "EVENT_DATE",
                "EVENT_TYPE",
            ]
        ).reset_index(
            drop=True
        )

        intros = sorted(
            set(
                g.loc[
                    g[
                        "EVENT_TYPE"
                    ].eq(
                        "INTRO"
                    ),
                    "EVENT_DATE"
                ]
            )
        )

        exclusions = sorted(
            set(
                g.loc[
                    g[
                        "EVENT_TYPE"
                    ].eq(
                        "EXCLUSION"
                    ),
                    "EVENT_DATE"
                ]
            )
        )


        # If we have exclusions but no introduction, historical start is
        # unknown. Do not assume the stock was always in F&O.
        if (
            len(
                intros
            ) == 0
            and
            len(
                exclusions
            ) > 0
        ):

            ambiguities.append(
                {
                    "COMPANY_ID":
                        company_id,

                    "ISSUE":
                        "EXCLUSION_WITHOUT_INTRODUCTION",

                    "INTRO_DATES":
                        "",

                    "EXCLUSION_DATES":
                        ";".join(
                            str(
                                x.date()
                            )
                            for x in exclusions
                        ),
                }
            )

            continue


        used_exclusions = set()


        for intro_idx, intro_date in enumerate(
            intros
        ):

            next_intro = (
                intros[
                    intro_idx
                    +
                    1
                ]
                if intro_idx
                +
                1
                <
                len(
                    intros
                )
                else None
            )

            eligible_exclusions = [
                x
                for x in exclusions
                if (
                    x
                    >
                    intro_date
                    and
                    x not in used_exclusions
                    and
                    (
                        next_intro is None
                        or
                        x
                        <=
                        next_intro
                    )
                )
            ]


            if len(
                eligible_exclusions
            ) > 1:

                ambiguities.append(
                    {
                        "COMPANY_ID":
                            company_id,

                        "ISSUE":
                            "MULTIPLE_EXCLUSIONS_FOR_ONE_INTRODUCTION",

                        "INTRO_DATES":
                            str(
                                intro_date.date()
                            ),

                        "EXCLUSION_DATES":
                            ";".join(
                                str(
                                    x.date()
                                )
                                for x in eligible_exclusions
                            ),
                    }
                )

                continue


            exclusion_date = (
                eligible_exclusions[
                    0
                ]
                if len(
                    eligible_exclusions
                ) == 1
                else pd.NaT
            )


            if pd.notna(
                exclusion_date
            ):

                used_exclusions.add(
                    exclusion_date
                )


            symbols_for_company = ";".join(
                sorted(
                    set(
                        g[
                            "TRACKER_SYMBOL"
                        ].dropna().astype(str)
                    )
                )
            )


            episodes.append(
                {
                    "COMPANY_ID":
                        company_id,

                    "TRACKER_SYMBOLS":
                        symbols_for_company,

                    "INTRODUCTION_DATE":
                        intro_date,

                    "EXCLUSION_DATE":
                        exclusion_date,

                    "ELIGIBILITY_RULE":
                        (
                            "INTRODUCTION_DATE <= formation_date "
                            "< EXCLUSION_DATE"
                        ),
                }
            )


        unused_exclusions = [
            x
            for x in exclusions
            if x not in used_exclusions
        ]


        # An exclusion before the first introduction proves the workbook has
        # history we cannot safely pair.
        for exclusion_date in unused_exclusions:

            if (
                len(
                    intros
                ) == 0
                or
                exclusion_date
                <=
                min(
                    intros
                )
            ):

                ambiguities.append(
                    {
                        "COMPANY_ID":
                            company_id,

                        "ISSUE":
                            "UNPAIRED_EXCLUSION",

                        "INTRO_DATES":
                            ";".join(
                                str(
                                    x.date()
                                )
                                for x in intros
                            ),

                        "EXCLUSION_DATES":
                            str(
                                exclusion_date.date()
                            ),
                    }
                )


    episodes_df = pd.DataFrame(
        episodes
    )

    ambiguity_df = pd.DataFrame(
        ambiguities
    )


    if not episodes_df.empty:

        episodes_df[
            "EPISODE_ID"
        ] = [
            f"FNOEP{i:04d}"
            for i in range(
                1,
                len(
                    episodes_df
                )
                +
                1
            )
        ]


    return (
        episodes_df,
        ambiguity_df
    )


# =============================================================================
# 8. EVALUATE F&O STATUS
# =============================================================================

def evaluate_fno(
    company_id,
    formation_date,
    episodes_df,
    company_event_ambiguities
):
    # If chronology for this company is ambiguous, do not guess.
    if company_id in company_event_ambiguities:

        return {
            "FNO_ELIGIBLE":
                pd.NA,

            "FNO_STATUS":
                "AMBIGUOUS_EVENT_CHRONOLOGY",

            "FNO_EPISODE_ID":
                pd.NA,

            "FNO_INTRODUCTION_DATE":
                pd.NaT,

            "FNO_EXCLUSION_DATE":
                pd.NaT,
        }


    if episodes_df.empty:

        return {
            "FNO_ELIGIBLE":
                pd.NA,

            "FNO_STATUS":
                "NO_PARSED_FNO_EPISODES",

            "FNO_EPISODE_ID":
                pd.NA,

            "FNO_INTRODUCTION_DATE":
                pd.NaT,

            "FNO_EXCLUSION_DATE":
                pd.NaT,
        }


    eps = episodes_df[
        episodes_df[
            "COMPANY_ID"
        ].eq(
            company_id
        )
    ].copy()


    if eps.empty:

        return {
            "FNO_ELIGIBLE":
                pd.NA,

            "FNO_STATUS":
                "NO_TRACKER_EPISODE_FOR_COMPANY",

            "FNO_EPISODE_ID":
                pd.NA,

            "FNO_INTRODUCTION_DATE":
                pd.NaT,

            "FNO_EXCLUSION_DATE":
                pd.NaT,
        }


    active = eps[
        eps[
            "INTRODUCTION_DATE"
        ].le(
            formation_date
        )
        &
        (
            eps[
                "EXCLUSION_DATE"
            ].isna()
            |
            eps[
                "EXCLUSION_DATE"
            ].gt(
                formation_date
            )
        )
    ]


    if len(
        active
    ) == 1:

        ep = active.iloc[0]

        return {
            "FNO_ELIGIBLE":
                True,

            "FNO_STATUS":
                "ELIGIBLE_TRACKER_CONFIRMED",

            "FNO_EPISODE_ID":
                ep.get(
                    "EPISODE_ID",
                    pd.NA
                ),

            "FNO_INTRODUCTION_DATE":
                ep[
                    "INTRODUCTION_DATE"
                ],

            "FNO_EXCLUSION_DATE":
                ep[
                    "EXCLUSION_DATE"
                ],
        }


    if len(
        active
    ) > 1:

        return {
            "FNO_ELIGIBLE":
                pd.NA,

            "FNO_STATUS":
                "AMBIGUOUS_OVERLAPPING_EPISODES",

            "FNO_EPISODE_ID":
                ";".join(
                    active[
                        "EPISODE_ID"
                    ].astype(str)
                ),

            "FNO_INTRODUCTION_DATE":
                pd.NaT,

            "FNO_EXCLUSION_DATE":
                pd.NaT,
        }


    # If the company's earliest known introduction is AFTER this formation
    # date, it is safely not eligible.
    earliest_intro = eps[
        "INTRODUCTION_DATE"
    ].min()


    if formation_date < earliest_intro:

        return {
            "FNO_ELIGIBLE":
                False,

            "FNO_STATUS":
                "NOT_YET_INTRODUCED",

            "FNO_EPISODE_ID":
                pd.NA,

            "FNO_INTRODUCTION_DATE":
                earliest_intro,

            "FNO_EXCLUSION_DATE":
                pd.NaT,
        }


    # If all known episodes ended before this date, safely not eligible.
    ended_eps = eps[
        eps[
            "EXCLUSION_DATE"
        ].notna()
        &
        eps[
            "EXCLUSION_DATE"
        ].le(
            formation_date
        )
    ]


    if (
        len(
            ended_eps
        ) == len(
            eps
        )
        and
        len(
            eps
        ) > 0
    ):

        return {
            "FNO_ELIGIBLE":
                False,

            "FNO_STATUS":
                "EXCLUDED_BEFORE_FORMATION",

            "FNO_EPISODE_ID":
                pd.NA,

            "FNO_INTRODUCTION_DATE":
                pd.NaT,

            "FNO_EXCLUSION_DATE":
                ended_eps[
                    "EXCLUSION_DATE"
                ].max(),
        }


    return {
        "FNO_ELIGIBLE":
            False,

        "FNO_STATUS":
            "NOT_ELIGIBLE_ON_FORMATION_DATE",

        "FNO_EPISODE_ID":
            pd.NA,

        "FNO_INTRODUCTION_DATE":
            pd.NaT,

        "FNO_EXCLUSION_DATE":
            pd.NaT,
    }


# =============================================================================
# 9. RUN WORKBOOK PARSER
# =============================================================================

print("=" * 110)
print("HISTORICAL F&O ELIGIBILITY + FINAL INVESTABLE UNIVERSE — ROBUST V2")
print("=" * 110)

tracker_path = locate_tracker()

print("\nOfficial frozen NSE tracker:")
print(
    tracker_path
)

tracker_hash = sha256_file(
    tracker_path
)


source_manifest = pd.DataFrame(
    [
        {
            "TRACKER_FILE":
                str(
                    tracker_path
                ),

            "TRACKER_AS_OF":
                TRACKER_AS_OF,

            "FILE_SIZE_BYTES":
                tracker_path.stat().st_size,

            "SHA256":
                tracker_hash,

            "RESEARCH_SYMBOLS_SEARCHED":
                len(
                    research_symbols
                ),
        }
    ]
)

source_manifest.to_csv(
    OUTPUT_DIR
    /
    "01_NSE_FNO_TRACKER_SOURCE_MANIFEST.csv",
    index=False,
    date_format="%Y-%m-%d"
)


print("\nScanning workbook directly for research-universe symbols...")

(
    symbol_occurrences,
    event_records,
    unresolved_workbook_rows,
    workbook_sheets,
) = scan_tracker_for_research_symbols(
    tracker_path
)


symbol_occurrences.to_csv(
    OUTPUT_DIR
    /
    "02_TRACKER_RESEARCH_SYMBOL_OCCURRENCES.csv",
    index=False
)


event_records.to_csv(
    OUTPUT_DIR
    /
    "03_TRACKER_FNO_EVENT_RECORDS.csv",
    index=False,
    date_format="%Y-%m-%d"
)


unresolved_workbook_rows.to_csv(
    OUTPUT_DIR
    /
    "04_TRACKER_UNRESOLVED_WORKBOOK_ROWS.csv",
    index=False
)


if symbol_occurrences.empty:

    raise RuntimeError(
        "\nThe workbook opened successfully, but none of the research-universe "
        "symbols were found anywhere in it.\n\n"
        "This suggests the downloaded file is not the expected tracker."
    )


if event_records.empty:

    raise RuntimeError(
        "\nResearch-universe symbols WERE found in the official workbook, "
        "but no introduction/exclusion dates could be resolved safely.\n\n"
        "No F&O eligibility was guessed.\n"
        "Please upload these two small files if this occurs:\n"
        "02_TRACKER_RESEARCH_SYMBOL_OCCURRENCES.csv\n"
        "04_TRACKER_UNRESOLVED_WORKBOOK_ROWS.csv"
    )


episodes, episode_ambiguities = build_episodes(
    event_records
)


episodes.to_csv(
    OUTPUT_DIR
    /
    "05_NSE_FNO_EPISODES_NORMALIZED.csv",
    index=False,
    date_format="%Y-%m-%d"
)


episode_ambiguities.to_csv(
    OUTPUT_DIR
    /
    "06_FNO_EPISODE_AMBIGUITIES.csv",
    index=False
)


ambiguous_companies = set()

if not episode_ambiguities.empty:

    ambiguous_companies.update(
        episode_ambiguities[
            "COMPANY_ID"
        ].dropna().astype(str)
    )


# =============================================================================
# 10. APPLY LIQUIDITY RULE
# =============================================================================

diagnostics[
    "LIQUIDITY_THRESHOLD_RUPEES"
] = LIQUIDITY_THRESHOLD_RUPEES

diagnostics[
    "LIQUIDITY_THRESHOLD_CRORE"
] = (
    LIQUIDITY_THRESHOLD_RUPEES
    /
    1e7
)

diagnostics[
    "LIQUIDITY_PASS"
] = (
    diagnostics[
        "MEDIAN_DAILY_TRADED_VALUE"
    ].notna()
    &
    diagnostics[
        "MEDIAN_DAILY_TRADED_VALUE"
    ].ge(
        LIQUIDITY_THRESHOLD_RUPEES
    )
)


# =============================================================================
# 11. EVALUATE F&O AT EVERY FORMATION DATE
# =============================================================================

fno_results = []

for _, row in diagnostics.iterrows():

    result = evaluate_fno(
        row[
            "COMPANY_ID"
        ],
        row[
            "FORMATION_DATE"
        ],
        episodes,
        ambiguous_companies
    )

    fno_results.append(
        result
    )


fno_results = pd.DataFrame(
    fno_results
)


diagnostics = pd.concat(
    [
        diagnostics.reset_index(
            drop=True
        ),
        fno_results.reset_index(
            drop=True
        ),
    ],
    axis=1
)


diagnostics[
    "FNO_ELIGIBLE"
] = diagnostics[
    "FNO_ELIGIBLE"
].astype(
    "boolean"
)


# =============================================================================
# 12. FINAL INVESTABLE
# =============================================================================

diagnostics[
    "FINAL_INVESTABLE"
] = (
    diagnostics[
        "PRE_LIQUIDITY_ELIGIBLE"
    ]
    &
    diagnostics[
        "LIQUIDITY_PASS"
    ]
    &
    diagnostics[
        "FNO_ELIGIBLE"
    ].fillna(
        False
    )
)


def final_status(row):

    reasons = []

    if not row[
        "PRE_LIQUIDITY_ELIGIBLE"
    ]:
        reasons.append(
            "FAILED_PRE_LIQUIDITY_RULES"
        )

    if not row[
        "LIQUIDITY_PASS"
    ]:
        reasons.append(
            "FAILED_10_CRORE_LIQUIDITY_RULE"
        )

    if pd.isna(
        row[
            "FNO_ELIGIBLE"
        ]
    ):
        reasons.append(
            "FNO_AMBIGUOUS"
        )

    elif not bool(
        row[
            "FNO_ELIGIBLE"
        ]
    ):
        reasons.append(
            "NOT_FNO_ELIGIBLE"
        )

    if not reasons:
        return "FINAL_INVESTABLE"

    return ";".join(
        reasons
    )


diagnostics[
    "FINAL_INVESTABLE_STATUS"
] = diagnostics.apply(
    final_status,
    axis=1
)


# =============================================================================
# 13. AUDITS AND OUTPUT TABLES
# =============================================================================

candidate_relevant = diagnostics[
    diagnostics[
        "PRE_LIQUIDITY_ELIGIBLE"
    ]
    &
    diagnostics[
        "LIQUIDITY_PASS"
    ]
].copy()


ambiguous_fno = candidate_relevant[
    candidate_relevant[
        "FNO_ELIGIBLE"
    ].isna()
].copy()


final_investable = diagnostics[
    diagnostics[
        "FINAL_INVESTABLE"
    ]
].copy()


final_investable = final_investable.sort_values(
    [
        "FORMATION_DATE",
        "COMPANY_ID",
    ]
).reset_index(
    drop=True
)


formation_counts = (
    diagnostics
    .groupby(
        [
            "FORMATION_DATE",
            "BLOCK_TYPE",
        ],
        as_index=False
    )
    .agg(
        INDEX_MEMBERS=(
            "IN_NIFTY_PHARMA_AT_FORMATION",
            "sum"
        ),

        PRE_LIQUIDITY_ELIGIBLE=(
            "PRE_LIQUIDITY_ELIGIBLE",
            "sum"
        ),

        PASS_10CR_LIQUIDITY=(
            "LIQUIDITY_PASS",
            lambda x:
                int(
                    x.sum()
                )
        ),

        FINAL_INVESTABLE=(
            "FINAL_INVESTABLE",
            "sum"
        ),
    )
)


# More precise candidate-stage counts.
formation_extra = []

for (
    formation_date,
    block_type
), grp in diagnostics.groupby(
    [
        "FORMATION_DATE",
        "BLOCK_TYPE",
    ]
):

    pre = grp[
        grp[
            "PRE_LIQUIDITY_ELIGIBLE"
        ]
    ]

    liq = pre[
        pre[
            "LIQUIDITY_PASS"
        ]
    ]

    formation_extra.append(
        {
            "FORMATION_DATE":
                formation_date,

            "BLOCK_TYPE":
                block_type,

            "PRE_LIQ_AND_LIQUID":
                len(
                    liq
                ),

            "CONFIRMED_FNO_TRUE_AFTER_LIQ":
                int(
                    liq[
                        "FNO_ELIGIBLE"
                    ].eq(
                        True
                    ).sum()
                ),

            "CONFIRMED_FNO_FALSE_AFTER_LIQ":
                int(
                    liq[
                        "FNO_ELIGIBLE"
                    ].eq(
                        False
                    ).sum()
                ),

            "AMBIGUOUS_FNO_AFTER_LIQ":
                int(
                    liq[
                        "FNO_ELIGIBLE"
                    ].isna().sum()
                ),
        }
    )


formation_extra = pd.DataFrame(
    formation_extra
)


formation_counts = formation_counts.merge(
    formation_extra,
    on=[
        "FORMATION_DATE",
        "BLOCK_TYPE",
    ],
    how="left",
    validate="one_to_one"
)


eligible_by_formation = (
    final_investable
    .groupby(
        [
            "FORMATION_DATE",
            "BLOCK_TYPE",
        ],
        as_index=False
    )
    .agg(
        N_FINAL_INVESTABLE=(
            "COMPANY_ID",
            "nunique"
        ),

        FINAL_INVESTABLE_COMPANIES=(
            "COMPANY_ID",
            lambda x:
                ";".join(
                    sorted(
                        set(
                            x.dropna().astype(str)
                        )
                    )
                )
        ),
    )
)


formation_master = diagnostics[
    [
        "FORMATION_DATE",
        "BLOCK_TYPE",
    ]
].drop_duplicates()


eligible_by_formation = formation_master.merge(
    eligible_by_formation,
    on=[
        "FORMATION_DATE",
        "BLOCK_TYPE",
    ],
    how="left"
)


eligible_by_formation[
    "N_FINAL_INVESTABLE"
] = eligible_by_formation[
    "N_FINAL_INVESTABLE"
].fillna(
    0
).astype(
    int
)


eligible_by_formation[
    "FINAL_INVESTABLE_COMPANIES"
] = eligible_by_formation[
    "FINAL_INVESTABLE_COMPANIES"
].fillna(
    ""
)


liquidity_exclusions = diagnostics[
    diagnostics[
        "PRE_LIQUIDITY_ELIGIBLE"
    ]
    &
    ~diagnostics[
        "LIQUIDITY_PASS"
    ]
].copy()


fno_exclusions = candidate_relevant[
    candidate_relevant[
        "FNO_ELIGIBLE"
    ].eq(
        False
    )
].copy()


# Coverage by research company.
coverage_records = []

for company_id in sorted(
    research_company_ids
):

    eps = (
        episodes[
            episodes[
                "COMPANY_ID"
            ].eq(
                company_id
            )
        ]
        if not episodes.empty
        else pd.DataFrame()
    )

    d = diagnostics[
        diagnostics[
            "COMPANY_ID"
        ].eq(
            company_id
        )
    ]

    relevant = d[
        d[
            "PRE_LIQUIDITY_ELIGIBLE"
        ]
        &
        d[
            "LIQUIDITY_PASS"
        ]
    ]

    coverage_records.append(
        {
            "COMPANY_ID":
                company_id,

            "SYMBOL_FOUND_IN_WORKBOOK":
                (
                    company_id
                    in set(
                        symbol_occurrences[
                            "COMPANY_ID"
                        ].dropna()
                    )
                ),

            "N_FNO_EPISODES":
                len(
                    eps
                ),

            "LIQUID_CANDIDATE_ROWS":
                len(
                    relevant
                ),

            "FNO_TRUE_ROWS":
                int(
                    relevant[
                        "FNO_ELIGIBLE"
                    ].eq(
                        True
                    ).sum()
                ),

            "FNO_FALSE_ROWS":
                int(
                    relevant[
                        "FNO_ELIGIBLE"
                    ].eq(
                        False
                    ).sum()
                ),

            "FNO_AMBIGUOUS_ROWS":
                int(
                    relevant[
                        "FNO_ELIGIBLE"
                    ].isna().sum()
                ),
        }
    )


company_coverage = pd.DataFrame(
    coverage_records
)


# =============================================================================
# 14. FINAL AUDIT STATUS
# =============================================================================

min_final_names = int(
    formation_counts[
        "FINAL_INVESTABLE"
    ].min()
)


if len(
    ambiguous_fno
) == 0:

    overall_status = "PASS"

else:

    overall_status = "PASS_WITH_REVIEW_ITEMS"


audit_summary = pd.DataFrame(
    [
        {
            "CHECK":
                "Overall status",

            "VALUE":
                overall_status,
        },
        {
            "CHECK":
                "Frozen official tracker date",

            "VALUE":
                str(
                    TRACKER_AS_OF.date()
                ),
        },
        {
            "CHECK":
                "Tracker SHA256",

            "VALUE":
                tracker_hash,
        },
        {
            "CHECK":
                "Workbook sheets",

            "VALUE":
                ";".join(
                    workbook_sheets
                ),
        },
        {
            "CHECK":
                "Research symbols searched",

            "VALUE":
                len(
                    research_symbols
                ),
        },
        {
            "CHECK":
                "Research-symbol workbook occurrences",

            "VALUE":
                len(
                    symbol_occurrences
                ),
        },
        {
            "CHECK":
                "Resolved F&O event records",

            "VALUE":
                len(
                    event_records
                ),
        },
        {
            "CHECK":
                "Normalized F&O episodes",

            "VALUE":
                len(
                    episodes
                ),
        },
        {
            "CHECK":
                "Episode chronology ambiguity rows",

            "VALUE":
                len(
                    episode_ambiguities
                ),
        },
        {
            "CHECK":
                "Stock x formation rows",

            "VALUE":
                len(
                    diagnostics
                ),
        },
        {
            "CHECK":
                "Frozen liquidity threshold crore",

            "VALUE":
                10.0,
        },
        {
            "CHECK":
                "Pre-liquidity eligible rows",

            "VALUE":
                int(
                    diagnostics[
                        "PRE_LIQUIDITY_ELIGIBLE"
                    ].sum()
                ),
        },
        {
            "CHECK":
                "Pre-liquidity rows passing liquidity",

            "VALUE":
                int(
                    (
                        diagnostics[
                            "PRE_LIQUIDITY_ELIGIBLE"
                        ]
                        &
                        diagnostics[
                            "LIQUIDITY_PASS"
                        ]
                    ).sum()
                ),
        },
        {
            "CHECK":
                "Ambiguous F&O rows after liquidity",

            "VALUE":
                len(
                    ambiguous_fno
                ),
        },
        {
            "CHECK":
                "Final investable stock-formation rows",

            "VALUE":
                len(
                    final_investable
                ),
        },
        {
            "CHECK":
                "Minimum final names on any formation date",

            "VALUE":
                min_final_names,
        },
    ]
)


# =============================================================================
# 15. SAVE
# =============================================================================

audit_summary.to_csv(
    OUTPUT_DIR
    /
    "00_FNO_INVESTABILITY_AUDIT_SUMMARY.csv",
    index=False
)


diagnostics.to_csv(
    OUTPUT_DIR
    /
    "07_STOCK_FORMATION_FINAL_INVESTABILITY.csv",
    index=False,
    date_format="%Y-%m-%d"
)


final_investable.to_csv(
    OUTPUT_DIR
    /
    "08_FINAL_INVESTABLE_STOCKS_LONG.csv",
    index=False,
    date_format="%Y-%m-%d"
)


eligible_by_formation.to_csv(
    OUTPUT_DIR
    /
    "09_FINAL_INVESTABLE_BY_FORMATION.csv",
    index=False,
    date_format="%Y-%m-%d"
)


formation_counts.to_csv(
    OUTPUT_DIR
    /
    "10_FORMATION_UNIVERSE_COUNTS.csv",
    index=False,
    date_format="%Y-%m-%d"
)


ambiguous_fno.to_csv(
    OUTPUT_DIR
    /
    "11_AMBIGUOUS_FNO_MATCHES.csv",
    index=False,
    date_format="%Y-%m-%d"
)


liquidity_exclusions.to_csv(
    OUTPUT_DIR
    /
    "12_LIQUIDITY_EXCLUSIONS.csv",
    index=False,
    date_format="%Y-%m-%d"
)


fno_exclusions.to_csv(
    OUTPUT_DIR
    /
    "13_FNO_EXCLUSIONS.csv",
    index=False,
    date_format="%Y-%m-%d"
)


company_coverage.to_csv(
    OUTPUT_DIR
    /
    "14_FNO_COMPANY_COVERAGE_AUDIT.csv",
    index=False
)


# =============================================================================
# 16. README
# =============================================================================

readme = f"""
NIFTY PHARMA — HISTORICAL F&O + FINAL INVESTABLE UNIVERSE — ROBUST V2

WHY V2
------
The NSE workbook layout did not match the single-row-header assumption in V1.

V2 does not depend on a specific worksheet name, header row, or column label.

It scans the official workbook directly for symbols belonging to our research
universe and classifies date cells from sheet names plus multi-row heading
context.

FROZEN LIQUIDITY RULE
---------------------
Median formation-window TOTAL_TRADED_VALUE >= Rs 10 crore.

F&O RULE
--------
For a confirmed F&O episode:

INTRODUCTION_DATE <= FORMATION_DATE < EXCLUSION_DATE

Blank EXCLUSION_DATE means the episode remains open.

AMBIGUITY RULE
--------------
The code never guesses.

Ambiguous candidate rows have FNO_ELIGIBLE missing and therefore
FINAL_INVESTABLE = False.

Official tracker:
{tracker_path}

Tracker SHA256:
{tracker_hash}

Overall status:
{overall_status}
"""

(
    OUTPUT_DIR
    /
    "README_FINAL_INVESTABLE_UNIVERSE_V2.txt"
).write_text(
    readme,
    encoding="utf-8"
)


# =============================================================================
# 17. CONSOLE OUTPUT
# =============================================================================

print("\n")
print("=" * 110)
print("ROBUST F&O PARSING + FINAL INVESTABLE UNIVERSE COMPLETE")
print("=" * 110)

print("\nAUDIT SUMMARY")
print("-" * 110)

print(
    audit_summary.to_string(
        index=False
    )
)


print("\nFORMATION COUNTS")
print("-" * 110)

print(
    formation_counts.to_string(
        index=False
    )
)


if len(
    ambiguous_fno
) > 0:

    print("\n")
    print("!" * 110)
    print(
        "SOME LIQUID CANDIDATES HAVE AMBIGUOUS F&O HISTORY."
    )
    print(
        "THEY WERE NOT TREATED AS INVESTABLE."
    )
    print("!" * 110)

    print(
        ambiguous_fno[
            [
                "FORMATION_DATE",
                "COMPANY_ID",
                "FNO_STATUS",
            ]
        ].to_string(
            index=False
        )
    )


print("\nOutputs:")
print(
    OUTPUT_DIR
)

print("\nUpload these next:")

for name in [
    "00_FNO_INVESTABILITY_AUDIT_SUMMARY.csv",
    "03_TRACKER_FNO_EVENT_RECORDS.csv",
    "05_NSE_FNO_EPISODES_NORMALIZED.csv",
    "06_FNO_EPISODE_AMBIGUITIES.csv",
    "09_FINAL_INVESTABLE_BY_FORMATION.csv",
    "10_FORMATION_UNIVERSE_COUNTS.csv",
    "11_AMBIGUOUS_FNO_MATCHES.csv",
    "12_LIQUIDITY_EXCLUSIONS.csv",
    "13_FNO_EXCLUSIONS.csv",
    "14_FNO_COMPANY_COVERAGE_AUDIT.csv",
]:

    print(
        OUTPUT_DIR
        /
        name
    )

print("\n")
print("=" * 110)
print("NO PAIR SELECTION OR P&L WAS USED.")
print("=" * 110)
