import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, date
import requests, zipfile, io, csv, hashlib, re, time, shutil, os
from openpyxl import load_workbook

# Config can be overridden by environment variables for testing.
PROJECT_ROOT = Path(os.environ.get('NIFTY_PROJECT_ROOT', r'C:\fin proj'))
FORMATION_DIAGNOSTICS_FILE = Path(os.environ.get('NIFTY_DIAGNOSTICS_FILE', str(PROJECT_ROOT / 'nse_pharma_formation_investability' / '02_STOCK_FORMATION_DIAGNOSTICS.csv')))
OUTPUT_DIR = Path(os.environ.get('NIFTY_FNO_OUTPUT_DIR', str(PROJECT_ROOT / 'nse_pharma_final_investable_universe_FINAL')))
RAW_TRACKER_DIR = OUTPUT_DIR / 'raw_nse_fno_tracker'
RAW_FO_DIR = OUTPUT_DIR / 'raw_nse_fo_bhavcopy'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
RAW_TRACKER_DIR.mkdir(parents=True, exist_ok=True)
RAW_FO_DIR.mkdir(parents=True, exist_ok=True)

LIQUIDITY_THRESHOLD_RUPEES = 100_000_000.0
UDIFF_START_DATE = pd.Timestamp('2024-07-08')
TRACKER_AS_OF = pd.Timestamp('2026-07-24')
TRACKER_FILENAME = 'FO_Stock_Introduction_and_Exclusion_Tracker_24-07-2026_20260724111707.xlsx'
TRACKER_LOCAL_FILE = RAW_TRACKER_DIR / TRACKER_FILENAME
TRACKER_URLS = [
    'https://nsearchives.nseindia.com/web/mediaattachment/2026-07/' + TRACKER_FILENAME,
    'https://nsearchives.nseindia.com//web/mediaattachment/2026-07/' + TRACKER_FILENAME,
]

ALIASES = {'CADILAHC':'ZYDUS','ZYDUSLIFE':'ZYDUS','AJANTAPHARM':'AJANTPHARM'}

def clean_symbol(v):
    if v is None or (isinstance(v,float) and np.isnan(v)):
        return None
    s = re.sub(r'\s+','',str(v).strip().upper())
    return None if s in {'','NAN','NONE','NA','N/A','-','--'} else s

def company_id_from_symbol(s):
    s=clean_symbol(s)
    return None if s is None else ALIASES.get(s,s)

def clean_col(v):
    return re.sub(r'[^A-Z0-9]+','_',str(v).strip().upper()).strip('_')

def parse_date(v):
    if v is None or (isinstance(v,float) and np.isnan(v)):
        return pd.NaT
    if isinstance(v,(pd.Timestamp,datetime,date)):
        return pd.Timestamp(v).normalize()
    t=str(v).strip()
    if not t: return pd.NaT
    t = re.sub(r',(?=\d)', ', ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    x=pd.to_datetime(t, errors='coerce')
    if pd.isna(x):
        x=pd.to_datetime(t, dayfirst=True, errors='coerce')
    return pd.NaT if pd.isna(x) else pd.Timestamp(x).normalize()

def to_bool(s):
    return s.astype(str).str.strip().str.upper().isin(['TRUE','1','YES','Y'])

def sha256_file(p):
    h=hashlib.sha256()
    with open(p,'rb') as f:
        for chunk in iter(lambda:f.read(1024*1024), b''):
            h.update(chunk)
    return h.hexdigest()

def valid_zip_bytes(b):
    if not isinstance(b,(bytes,bytearray)) or len(b)<200 or b[:2]!=b'PK': return False
    try:
        with zipfile.ZipFile(io.BytesIO(b)) as z: return z.testzip() is None
    except Exception: return False

def robust_download(urls, target, what):
    if target.exists():
        b=target.read_bytes()
        if valid_zip_bytes(b) if target.suffix.lower()=='.zip' else len(b)>1000:
            return 'CACHE'
        stamp=datetime.now().strftime('%Y%m%d_%H%M%S')
        corrupt=target.with_name(target.name+f'.corrupt_{stamp}')
        target.rename(corrupt)
    sess=requests.Session()
    headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/151 Safari/537.36','Accept':'*/*','Connection':'keep-alive'}
    errs=[]
    for attempt in range(1,4):
        for url in urls:
            try:
                r=sess.get(url,headers=headers,timeout=60,allow_redirects=True)
                good=(r.status_code==200 and (valid_zip_bytes(r.content) if target.suffix.lower()=='.zip' else len(r.content)>1000))
                if good:
                    target.write_bytes(r.content); return url
                errs.append(f'{url} HTTP {r.status_code} bytes={len(r.content)}')
            except Exception as e:
                errs.append(f'{url}: {type(e).__name__}: {e}')
        time.sleep(attempt)
    raise RuntimeError(f'Could not download {what}.\n'+'\n'.join(errs))

def locate_or_download_tracker():
    # Test override or existing project cache
    override=os.environ.get('NIFTY_TRACKER_FILE')
    if override and Path(override).exists():
        src=Path(override)
        if src.resolve()!=TRACKER_LOCAL_FILE.resolve(): shutil.copy2(src,TRACKER_LOCAL_FILE)
        return TRACKER_LOCAL_FILE,'LOCAL_OVERRIDE'
    if TRACKER_LOCAL_FILE.exists():
        try:
            with zipfile.ZipFile(TRACKER_LOCAL_FILE) as z:
                if '[Content_Types].xml' in z.namelist():
                    return TRACKER_LOCAL_FILE,'CACHE'
        except Exception:
            pass
        stamp=datetime.now().strftime('%Y%m%d_%H%M%S')
        TRACKER_LOCAL_FILE.rename(TRACKER_LOCAL_FILE.with_name(TRACKER_LOCAL_FILE.name+f'.corrupt_{stamp}'))
    matches=list(PROJECT_ROOT.rglob(TRACKER_FILENAME)) if PROJECT_ROOT.exists() else []
    if matches:
        shutil.copy2(matches[0],TRACKER_LOCAL_FILE); return TRACKER_LOCAL_FILE,'PROJECT_COPY'
    source=robust_download(TRACKER_URLS,TRACKER_LOCAL_FILE,'official NSE F&O tracker')
    return TRACKER_LOCAL_FILE,source

def merged_cell_value(ws,row,col):
    cell=ws.cell(row=row,column=col)
    if cell.value is not None: return cell.value
    for rng in ws.merged_cells.ranges:
        if rng.min_row<=row<=rng.max_row and rng.min_col<=col<=rng.max_col:
            return ws.cell(rng.min_row,rng.min_col).value
    return None

def parse_tracker(path):
    wb=load_workbook(path,data_only=True,read_only=False)
    rec=[]
    for sheet_name,event_type in [('Introduction','INTRODUCTION'),('Exclusion','EXCLUSION')]:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f'Tracker missing sheet {sheet_name}. Actual sheets: {wb.sheetnames}')
        ws=wb[sheet_name]
        headers={clean_col(ws.cell(2,c).value):c for c in range(1,ws.max_column+1) if ws.cell(2,c).value is not None}
        for needed in ['SR_NO','SYMBOL','SECURITY_NAME','EFFECTIVE_DATE','CIRCULAR_LINK']:
            if needed not in headers: raise RuntimeError(f'{sheet_name} missing header {needed}: {headers}')
        for r in range(3,ws.max_row+1):
            sr=ws.cell(r,headers['SR_NO']).value
            try: int(sr)
            except Exception: continue
            sym=clean_symbol(ws.cell(r,headers['SYMBOL']).value)
            if not sym: continue
            eff=merged_cell_value(ws,r,headers['EFFECTIVE_DATE'])
            link=merged_cell_value(ws,r,headers['CIRCULAR_LINK'])
            eff=parse_date(eff)
            if pd.isna(eff): raise RuntimeError(f'Could not resolve merged Effective Date in {sheet_name} row {r} symbol {sym}')
            rec.append({'EVENT_TYPE':event_type,'TRACKER_SYMBOL':sym,'COMPANY_ID':company_id_from_symbol(sym),'SECURITY_NAME':ws.cell(r,headers['SECURITY_NAME']).value,'EFFECTIVE_DATE':eff,'CIRCULAR_LINK':link,'SOURCE_SHEET':sheet_name,'SOURCE_ROW':r})
    df=pd.DataFrame(rec).sort_values(['EFFECTIVE_DATE','EVENT_TYPE','TRACKER_SYMBOL']).reset_index(drop=True)
    return df

def fo_urls(dt):
    dt=pd.Timestamp(dt)
    if dt < UDIFF_START_DATE:
        y=dt.strftime('%Y'); mon=dt.strftime('%b').upper(); d=dt.strftime('%d'); fname=f'fo{d}{mon}{y}bhav.csv.zip'
        urls=[f'https://nsearchives.nseindia.com/content/historical/DERIVATIVES/{y}/{mon}/{fname}',f'https://archives.nseindia.com/content/historical/DERIVATIVES/{y}/{mon}/{fname}',f'https://www1.nseindia.com/content/historical/DERIVATIVES/{y}/{mon}/{fname}']
        schema='LEGACY'; original=fname
    else:
        ymd=dt.strftime('%Y%m%d'); fname=f'BhavCopy_NSE_FO_0_0_0_{ymd}_F_0000.csv.zip'
        urls=[f'https://nsearchives.nseindia.com/content/fo/{fname}',f'https://archives.nseindia.com/content/fo/{fname}']
        schema='UDIFF'; original=fname
    return urls,schema,original

def decode_csv_bytes(b):
    for enc in ['utf-8-sig','utf-8','cp1252','latin-1']:
        try: return b.decode(enc)
        except UnicodeDecodeError: pass
    return b.decode('latin-1',errors='replace')

def parse_fo_zip(path, expected_schema):
    with zipfile.ZipFile(path) as z:
        csv_names=[n for n in z.namelist() if n.lower().endswith('.csv')]
        if not csv_names: raise RuntimeError(f'No CSV inside {path}')
        last_headers=None
        for name in csv_names:
            txt=decode_csv_bytes(z.read(name))
            reader=csv.DictReader(io.StringIO(txt))
            if not reader.fieldnames: continue
            colmap={clean_col(x):x for x in reader.fieldnames}
            last_headers=list(reader.fieldnames)
            if 'INSTRUMENT' in colmap and 'SYMBOL' in colmap:
                eligible=set(); rows=0
                for row in reader:
                    if str(row.get(colmap['INSTRUMENT'],'')).strip().upper()=='FUTSTK':
                        sym=clean_symbol(row.get(colmap['SYMBOL']))
                        if sym: eligible.add(company_id_from_symbol(sym)); rows+=1
                if not eligible: raise RuntimeError(f'Legacy file parsed but no FUTSTK symbols: {path}')
                return eligible,'LEGACY_FUTSTK',rows,name,last_headers
            if 'FININSTRMTP' in colmap and 'TCKRSYMB' in colmap:
                eligible=set(); rows=0
                for row in reader:
                    if str(row.get(colmap['FININSTRMTP'],'')).strip().upper()=='STF':
                        sym=clean_symbol(row.get(colmap['TCKRSYMB']))
                        if sym: eligible.add(company_id_from_symbol(sym)); rows+=1
                if not eligible: raise RuntimeError(f'UDiFF file parsed but no STF symbols: {path}')
                return eligible,'UDIFF_STF',rows,name,last_headers
        raise RuntimeError(f'No recognized F&O schema in {path}. Last headers={last_headers}')

def get_fo_snapshot(dt):
    urls,expected_schema,original=fo_urls(dt)
    target=RAW_FO_DIR / original
    # Test override: if exact expected file exists in externally provided cache, copy it.
    ext=os.environ.get('NIFTY_FO_CACHE_DIR')
    if not target.exists() and ext:
        src=Path(ext)/original
        if src.exists(): shutil.copy2(src,target)
    source=robust_download(urls,target,f'NSE F&O bhavcopy for {pd.Timestamp(dt).date()}')
    eligible,parsed_schema,contract_rows,member_name,headers=parse_fo_zip(target,expected_schema)
    return {'FORMATION_DATE':pd.Timestamp(dt),'EXPECTED_SCHEMA':expected_schema,'PARSED_SCHEMA':parsed_schema,'ZIP_FILE':str(target),'ZIP_MEMBER':member_name,'DOWNLOAD_SOURCE':source,'ZIP_SHA256':sha256_file(target),'STOCK_FUTURE_CONTRACT_ROWS':contract_rows,'UNIQUE_FNO_COMPANIES':len(eligible),'ELIGIBLE_SET':eligible,'HEADERS':'|'.join(headers)}

def tracker_expectation(tracker_events,company_id,formation_date):
    sub=tracker_events[(tracker_events.COMPANY_ID==company_id)&(tracker_events.EFFECTIVE_DATE<=formation_date)].sort_values(['EFFECTIVE_DATE','SOURCE_ROW'])
    if sub.empty: return pd.NA,pd.NaT,pd.NA
    last=sub.iloc[-1]
    exp=True if last.EVENT_TYPE=='INTRODUCTION' else False
    return exp,last.EFFECTIVE_DATE,last.EVENT_TYPE

print('='*110) ; print('NIFTY PHARMA — FINAL HISTORICAL F&O INVESTABILITY (V3)'); print('='*110)

# Load diagnostics
if not FORMATION_DIAGNOSTICS_FILE.exists(): raise FileNotFoundError(FORMATION_DIAGNOSTICS_FILE)
df=pd.read_csv(FORMATION_DIAGNOSTICS_FILE,low_memory=False)
df.columns=[clean_col(c) for c in df.columns]
req={'FORMATION_DATE','COMPANY_ID','BLOCK_TYPE','IN_NIFTY_PHARMA_AT_FORMATION','PRE_LIQUIDITY_ELIGIBLE','MEDIAN_DAILY_TRADED_VALUE'}
missing=req-set(df.columns)
if missing: raise RuntimeError(f'Diagnostics missing columns: {sorted(missing)}')
df['FORMATION_DATE']=pd.to_datetime(df['FORMATION_DATE'],errors='raise').dt.normalize()
df['COMPANY_ID']=df['COMPANY_ID'].map(clean_symbol)
df['PRE_LIQUIDITY_ELIGIBLE']=to_bool(df['PRE_LIQUIDITY_ELIGIBLE'])
df['IN_NIFTY_PHARMA_AT_FORMATION']=to_bool(df['IN_NIFTY_PHARMA_AT_FORMATION'])
df['MEDIAN_DAILY_TRADED_VALUE']=pd.to_numeric(df['MEDIAN_DAILY_TRADED_VALUE'],errors='coerce')
if df[['FORMATION_DATE','COMPANY_ID']].duplicated().any(): raise RuntimeError('Duplicate FORMATION_DATE+COMPANY_ID')

# Tracker: archive and parse actual merged layout
tracker_path,tracker_source=locate_or_download_tracker()
tracker_events=parse_tracker(tracker_path)
tracker_events.to_csv(OUTPUT_DIR/'02_NSE_FNO_TRACKER_EVENTS_PARSED.csv',index=False,date_format='%Y-%m-%d')
pd.DataFrame([{'TRACKER_FILE':str(tracker_path),'TRACKER_SOURCE':tracker_source,'TRACKER_AS_OF':TRACKER_AS_OF,'SHA256':sha256_file(tracker_path),'EVENT_ROWS':len(tracker_events)}]).to_csv(OUTPUT_DIR/'01_NSE_FNO_TRACKER_SOURCE_MANIFEST.csv',index=False,date_format='%Y-%m-%d')

# Direct official daily F&O evidence for every formation date
snapshots=[]; sets={}
for dt in sorted(df['FORMATION_DATE'].unique()):
    print(f'F&O snapshot {pd.Timestamp(dt).date()} ...')
    s=get_fo_snapshot(dt)
    sets[pd.Timestamp(dt)]=s.pop('ELIGIBLE_SET')
    snapshots.append(s)
snapshot_df=pd.DataFrame(snapshots)
snapshot_df.to_csv(OUTPUT_DIR/'03_FORMATION_DATE_FNO_BHAVCOPY_AUDIT.csv',index=False,date_format='%Y-%m-%d')

# Apply rules
df['LIQUIDITY_THRESHOLD_RUPEES']=LIQUIDITY_THRESHOLD_RUPEES
df['LIQUIDITY_THRESHOLD_CRORE']=10.0
df['LIQUIDITY_PASS']=df['MEDIAN_DAILY_TRADED_VALUE'].ge(LIQUIDITY_THRESHOLD_RUPEES)&df['MEDIAN_DAILY_TRADED_VALUE'].notna()
df['FNO_ELIGIBLE']=[row.COMPANY_ID in sets[row.FORMATION_DATE] for row in df.itertuples(index=False)]
df['FNO_EVIDENCE']='OFFICIAL_NSE_DAILY_FO_BHAVCOPY_STOCK_FUTURE_PRESENT'

# Tracker cross-check only where tracker has a prior recorded state change.
tr_exp=[]; tr_date=[]; tr_event=[]
for row in df.itertuples(index=False):
    e,d,t=tracker_expectation(tracker_events,row.COMPANY_ID,row.FORMATION_DATE)
    tr_exp.append(e); tr_date.append(d); tr_event.append(t)
df['TRACKER_EXPECTED_FNO_STATE']=pd.Series(tr_exp,dtype='boolean')
df['TRACKER_LAST_EVENT_DATE']=tr_date
df['TRACKER_LAST_EVENT_TYPE']=tr_event
df['TRACKER_DIRECT_CONSISTENT']=pd.Series([pd.NA if pd.isna(e) else bool(e)==bool(f) for e,f in zip(tr_exp,df['FNO_ELIGIBLE'])],dtype='boolean')

df['FINAL_INVESTABLE']=df['PRE_LIQUIDITY_ELIGIBLE'] & df['LIQUIDITY_PASS'] & df['FNO_ELIGIBLE']

def status(r):
    x=[]
    if not r.PRE_LIQUIDITY_ELIGIBLE: x.append('FAILED_PRE_LIQUIDITY_RULES')
    if not r.LIQUIDITY_PASS: x.append('FAILED_10_CRORE_LIQUIDITY_RULE')
    if not r.FNO_ELIGIBLE: x.append('NO_STOCK_FUTURE_ON_FORMATION_DATE')
    return 'FINAL_INVESTABLE' if not x else ';'.join(x)
df['FINAL_INVESTABLE_STATUS']=[status(r) for r in df.itertuples(index=False)]

# Audit tables
fno_snap_rows=[]
research=set(df.COMPANY_ID.dropna())
for dt,eligible in sets.items():
    for cid in sorted(research):
        fno_snap_rows.append({'FORMATION_DATE':dt,'COMPANY_ID':cid,'FNO_ELIGIBLE':cid in eligible})
pd.DataFrame(fno_snap_rows).to_csv(OUTPUT_DIR/'04_RESEARCH_UNIVERSE_FNO_SNAPSHOT.csv',index=False,date_format='%Y-%m-%d')

df.to_csv(OUTPUT_DIR/'05_STOCK_FORMATION_FINAL_INVESTABILITY.csv',index=False,date_format='%Y-%m-%d')
final_long=df[df.FINAL_INVESTABLE].copy(); final_long.to_csv(OUTPUT_DIR/'06_FINAL_INVESTABLE_STOCKS_LONG.csv',index=False,date_format='%Y-%m-%d')

forms=df[['FORMATION_DATE','BLOCK_TYPE']].drop_duplicates().sort_values('FORMATION_DATE')
agg=final_long.groupby(['FORMATION_DATE','BLOCK_TYPE']).agg(N_FINAL_INVESTABLE=('COMPANY_ID','nunique'),FINAL_INVESTABLE_COMPANIES=('COMPANY_ID',lambda x:';'.join(sorted(set(x))))).reset_index()
by_form=forms.merge(agg,on=['FORMATION_DATE','BLOCK_TYPE'],how='left'); by_form['N_FINAL_INVESTABLE']=by_form['N_FINAL_INVESTABLE'].fillna(0).astype(int); by_form['FINAL_INVESTABLE_COMPANIES']=by_form['FINAL_INVESTABLE_COMPANIES'].fillna('')
by_form.to_csv(OUTPUT_DIR/'07_FINAL_INVESTABLE_BY_FORMATION.csv',index=False,date_format='%Y-%m-%d')

counts=[]
for (dt,bt),g in df.groupby(['FORMATION_DATE','BLOCK_TYPE']):
    pre=g[g.PRE_LIQUIDITY_ELIGIBLE]; liq=pre[pre.LIQUIDITY_PASS]
    counts.append({'FORMATION_DATE':dt,'BLOCK_TYPE':bt,'INDEX_MEMBERS':int(g.IN_NIFTY_PHARMA_AT_FORMATION.sum()),'PRE_LIQUIDITY_ELIGIBLE':len(pre),'PASS_10CR_LIQUIDITY':len(liq),'FNO_ELIGIBLE_AFTER_LIQUIDITY':int(liq.FNO_ELIGIBLE.sum()),'FINAL_INVESTABLE':int(g.FINAL_INVESTABLE.sum())})
counts_df=pd.DataFrame(counts).sort_values('FORMATION_DATE'); counts_df.to_csv(OUTPUT_DIR/'08_FORMATION_UNIVERSE_COUNTS.csv',index=False,date_format='%Y-%m-%d')

tracker_review=df[df.TRACKER_DIRECT_CONSISTENT.eq(False)].copy(); tracker_review.to_csv(OUTPUT_DIR/'09_TRACKER_CROSSCHECK_REVIEW.csv',index=False,date_format='%Y-%m-%d')
df[df.PRE_LIQUIDITY_ELIGIBLE & ~df.LIQUIDITY_PASS].to_csv(OUTPUT_DIR/'10_LIQUIDITY_EXCLUSIONS.csv',index=False,date_format='%Y-%m-%d')
df[df.PRE_LIQUIDITY_ELIGIBLE & df.LIQUIDITY_PASS & ~df.FNO_ELIGIBLE].to_csv(OUTPUT_DIR/'11_FNO_EXCLUSIONS.csv',index=False,date_format='%Y-%m-%d')

coverage=[]
for cid,g in df.groupby('COMPANY_ID'):
    rel=g[g.PRE_LIQUIDITY_ELIGIBLE & g.LIQUIDITY_PASS]
    coverage.append({'COMPANY_ID':cid,'FORMATION_ROWS':len(g),'LIQUID_CANDIDATE_ROWS':len(rel),'FNO_TRUE_ROWS':int(rel.FNO_ELIGIBLE.sum()),'FNO_FALSE_ROWS':int((~rel.FNO_ELIGIBLE).sum()),'TRACKER_EVENTS':int((tracker_events.COMPANY_ID==cid).sum())})
pd.DataFrame(coverage).to_csv(OUTPUT_DIR/'12_FNO_COMPANY_COVERAGE_AUDIT.csv',index=False)

min_names=int(counts_df.FINAL_INVESTABLE.min())
status='PASS' if len(tracker_review)==0 and min_names>=2 else 'PASS_WITH_REVIEW_ITEMS'
audit=pd.DataFrame([
 {'CHECK':'Overall status','VALUE':status},
 {'CHECK':'Formation dates evaluated','VALUE':df.FORMATION_DATE.nunique()},
 {'CHECK':'Official NSE F&O bhavcopies parsed','VALUE':len(snapshot_df)},
 {'CHECK':'Bhavcopy download/parse failures','VALUE':0},
 {'CHECK':'Frozen liquidity threshold crore','VALUE':10.0},
 {'CHECK':'Pre-liquidity eligible rows','VALUE':int(df.PRE_LIQUIDITY_ELIGIBLE.sum())},
 {'CHECK':'Rows passing liquidity after pre-liquidity screen','VALUE':int((df.PRE_LIQUIDITY_ELIGIBLE & df.LIQUIDITY_PASS).sum())},
 {'CHECK':'Final investable stock-formation rows','VALUE':int(df.FINAL_INVESTABLE.sum())},
 {'CHECK':'Minimum final investable names per formation','VALUE':min_names},
 {'CHECK':'Tracker events parsed from merged workbook','VALUE':len(tracker_events)},
 {'CHECK':'Tracker/direct cross-check mismatches','VALUE':len(tracker_review)},
 {'CHECK':'Tracker role','VALUE':'AUDIT ONLY; historical F&O eligibility comes from daily official NSE F&O bhavcopy'},
])
audit.to_csv(OUTPUT_DIR/'00_FNO_INVESTABILITY_AUDIT_SUMMARY.csv',index=False)

(OUTPUT_DIR/'README_FINAL_FNO_UNIVERSE.txt').write_text(f'''NIFTY PHARMA FINAL F&O INVESTABILITY\n\nAuthoritative historical F&O test: stock future contract present in the official NSE daily F&O bhavcopy on each formation date.\nLegacy format: INSTRUMENT == FUTSTK.\nUDiFF format (from 08-Jul-2024): FinInstrmTp == STF.\n\nThe 24-Jul-2026 Introduction/Exclusion Tracker is archived and parsed with merged-cell handling, but is used only as a cross-check because it does not contain introduction history for legacy F&O stocks going back to 2017.\n\nFinal investable = PRE_LIQUIDITY_ELIGIBLE AND median daily traded value >= Rs 10 crore AND FNO_ELIGIBLE.\n\nOverall status: {status}\n''',encoding='utf-8')

print('\n'+audit.to_string(index=False))
print('\n'+counts_df.to_string(index=False))
print(f'\nOutputs: {OUTPUT_DIR}')
